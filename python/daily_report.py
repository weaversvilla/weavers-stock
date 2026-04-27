#!/usr/bin/env python3
"""
Weavers Villa — Daily Stock Intelligence
Runs via GitHub Actions at 6am IST
1. Fetches all data from Baselinker
2. Fetches Global sales from Amazon SP-API (US/UK/UAE)
3. Builds bundle map → remaps bundle sales to component SKUs
4. Calculates velocity + reorder quantities
5. Saves report to report_data.json
6. Uploads to GitHub (Vercel serves it instantly)
7. Sends HTML email to weavers.villa@gmail.com
"""

import json
import csv
import os
import time
import socket
import smtplib
import urllib.request
import urllib.error
import urllib.parse
import base64
import hmac
import hashlib
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

# ─── CONFIG ──────────────────────────────────────────────────────────────────
BL_TOKEN       = os.environ.get("BL_TOKEN",       "")
GITHUB_TOKEN   = os.environ.get("GH_PAT",         os.environ.get("GITHUB_TOKEN", ""))
GMAIL_APP_PASS = os.environ.get("GMAIL_APP_PASS", "")

# Amazon SP-API — Global Sales
SP_CLIENT_ID     = os.environ.get("SP_CLIENT_ID",     "")
SP_CLIENT_SECRET = os.environ.get("SP_CLIENT_SECRET", "")
SP_REFRESH_US    = os.environ.get("SP_REFRESH_US",    "")
SP_REFRESH_UK    = os.environ.get("SP_REFRESH_UK",    "")
SP_REFRESH_UAE   = os.environ.get("SP_REFRESH_UAE",   "")

MARKETPLACES = {
    "US":  {
        "marketplace_id": "ATVPDKIKX0DER",
        "endpoint":       "https://sellingpartnerapi-na.amazon.com",
        "region":         "us-east-1",
        "refresh_token":  SP_REFRESH_US,
    },
    "UK":  {
        "marketplace_id": "A1F83G8C2ARO7P",
        "endpoint":       "https://sellingpartnerapi-eu.amazon.com",
        "region":         "eu-west-1",
        "refresh_token":  SP_REFRESH_UK,
    },
    "UAE": {
        "marketplace_id": "A2VIGQ35RCS4UG",
        "endpoint":       "https://sellingpartnerapi-eu.amazon.com",
        "region":         "eu-west-1",
        "refresh_token":  SP_REFRESH_UAE,
    },
}

# Excel mapping file path in repo
GLOBAL_SKU_MAP_FILE = Path(__file__).parent.parent / "data" / "global_sku_mapping.xlsx"

GITHUB_REPO    = "weaversvilla/weavers-stock"
GITHUB_FILE    = "public/report_data.json"
GMAIL_FROM     = "weavers.villa@gmail.com"
GMAIL_TO       = "weavers.villa@gmail.com"

INVENTORY_ID   = 1257
WAREHOUSE_ID   = 9001890
TARGET_DAYS    = 90
SEASONAL_MULT  = 1.0

CANCELLED_STATUS_ID = 12225

EXCLUDED_SKUS = {
    "CHAIR-POPPY-SET-4", "CHAIR-POPPY-SET-6",
    "CHAIR-PRISM-SET-4", "CHAIR-PRISM-SET-6",
    "HT-RIBBON-SET-12",  "HT-RIBBON-SET-24",
    "HT-RIBBON-SET-4",   "HT-RIBBON-SET-6",
}

SNAPSHOT_CSV      = Path(__file__).parent / "velocity_history.csv"
VALUE_HISTORY_CSV = Path(__file__).parent / "stock_value_history.csv"

PLATFORM_MAP = {
    "amazon":         [443],
    "amazon_vendor":  [445, 537],
    "flipkart":       [1414],
    "myntra":         [1492],
    "ajio":           [1548],
    "others":         [1636, 1601, 9000337],
}

WEIGHTS = {"d7": 0.40, "d15": 0.30, "d30": 0.20, "d90": 0.10}
# ─────────────────────────────────────────────────────────────────────────────

BL_URL = "https://api.baselinker.com/connector.php"
socket.setdefaulttimeout(60)

def bl_call(method, params={}):
    """Make a Baselinker API call with rate limit and timeout handling."""
    data = urllib.parse.urlencode({
        "token": BL_TOKEN,
        "method": method,
        "parameters": json.dumps(params)
    }).encode()

    for attempt in range(5):
        try:
            req = urllib.request.Request(BL_URL, data=data, method="POST")
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = json.loads(resp.read().decode())

            if result.get("status") == "SUCCESS":
                return result

            err = result.get("error_message", "")
            if "limit exceeded" in err.lower() or "blocked" in err.lower():
                print(f"  Rate limited. Waiting 70 seconds...")
                time.sleep(70)
                continue

            raise Exception(f"BL API error [{method}]: {err}")

        except TimeoutError as e:
            print(f"  Timeout on attempt {attempt+1}, retrying in 5s...")
            time.sleep(5)
            continue
        except urllib.error.URLError as e:
            print(f"  Network error attempt {attempt+1}: {e}")
            time.sleep(5)

    raise Exception(f"BL API [{method}] failed after 5 attempts")

def days_ago(days):
    return int(time.time()) - days * 86400

def get_platform(source_id):
    for name, ids in PLATFORM_MAP.items():
        if source_id in ids:
            return name
    return "others"

# ── Amazon SP-API — Global Sales ──────────────────────────────────────────────

def sp_get_access_token(refresh_token):
    """Exchange refresh token for access token."""
    data = urllib.parse.urlencode({
        "grant_type":    "refresh_token",
        "refresh_token": refresh_token,
        "client_id":     SP_CLIENT_ID,
        "client_secret": SP_CLIENT_SECRET,
    }).encode()
    req = urllib.request.Request(
        "https://api.amazon.com/auth/o2/token",
        data=data,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        result = json.loads(resp.read().decode())
    return result["access_token"]

def sp_get_orders(marketplace, days=90):
    """Fetch orders from SP-API for a marketplace."""
    if not marketplace["refresh_token"]:
        return []

    try:
        access_token = sp_get_access_token(marketplace["refresh_token"])
    except Exception as e:
        print(f"    Token error: {e}")
        return []

    created_after = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")
    endpoint      = marketplace["endpoint"]
    market_id     = marketplace["marketplace_id"]

    all_orders = []
    next_token = None

    while True:
        params = {
            "MarketplaceIds": market_id,
            "CreatedAfter":   created_after,
            "OrderStatuses":  "Shipped,Unshipped,PartiallyShipped,Pending,InvoiceUnconfirmed",
        }
        if next_token:
            params = {"NextToken": next_token, "MarketplaceIds": market_id}

        url = f"{endpoint}/orders/v0/orders?" + urllib.parse.urlencode(params)
        req = urllib.request.Request(url, headers={
            "x-amz-access-token": access_token,
            "Content-Type":       "application/json",
        })

        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode())
        except urllib.error.HTTPError as e:
            if e.code == 429:
                print(f"    SP-API rate limited, waiting 60s...")
                time.sleep(60)
                continue
            print(f"    SP-API orders error: {e}")
            break
        except Exception as e:
            print(f"    SP-API orders error: {e}")
            break

        orders = data.get("payload", {}).get("Orders", [])
        all_orders.extend(orders)

        next_token = data.get("payload", {}).get("NextToken")
        if not next_token:
            break
        time.sleep(1.0)  # SP-API Orders list: max 1 req/sec

    return all_orders, access_token  # Return access_token to reuse

def sp_get_order_items(marketplace, order_id, access_token):
    """Fetch line items for a single SP-API order with retry on rate limit."""
    endpoint = marketplace["endpoint"]
    url = f"{endpoint}/orders/v0/orders/{order_id}/orderItems"
    req = urllib.request.Request(url, headers={
        "x-amz-access-token": access_token,
        "Content-Type":       "application/json",
    })
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode())
            return data.get("payload", {}).get("OrderItems", [])
        except urllib.error.HTTPError as e:
            if e.code == 429:
                wait = 2 ** attempt
                time.sleep(wait)
                continue
            return []
        except:
            return []
    return []

def load_global_sku_mapping():
    """Load global_sku → master_sku mapping from Excel file."""
    if not GLOBAL_SKU_MAP_FILE.exists():
        print(f"  ⚠️ Global SKU mapping file not found: {GLOBAL_SKU_MAP_FILE}")
        return {}

    try:
        # Read Excel without openpyxl dependency using csv fallback
        # Try openpyxl first
        try:
            import openpyxl
            wb = openpyxl.load_workbook(GLOBAL_SKU_MAP_FILE)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(h).lower().strip() if h else "" for h in rows[0]]
            g_idx = headers.index("global_sku")
            m_idx = headers.index("master_sku")
            mapping = {}
            for row in rows[1:]:
                if row[g_idx] and row[m_idx]:
                    mapping[str(row[g_idx]).strip()] = str(row[m_idx]).strip()
            print(f"  Global SKU mapping loaded: {len(mapping)} entries.")
            return mapping
        except ImportError:
            print("  openpyxl not installed, installing...")
            import subprocess
            subprocess.run(["pip", "install", "openpyxl", "--break-system-packages", "-q"])
            import openpyxl
            wb = openpyxl.load_workbook(GLOBAL_SKU_MAP_FILE)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(h).lower().strip() if h else "" for h in rows[0]]
            g_idx = headers.index("global_sku")
            m_idx = headers.index("master_sku")
            mapping = {}
            for row in rows[1:]:
                if row[g_idx] and row[m_idx]:
                    mapping[str(row[g_idx]).strip()] = str(row[m_idx]).strip()
            print(f"  Global SKU mapping loaded: {len(mapping)} entries.")
            return mapping
    except Exception as e:
        print(f"  ⚠️ Error loading SKU mapping: {e}")
        return {}

def get_global_sales(days=90):
    """
    Fetch sales from all global marketplaces (US/UK/UAE).
    Returns:
      - sales_by_master: {master_sku: {d7, d15, d30, d90}} units
      - unmapped_skus: set of global SKUs not in mapping file
    """
    if not SP_CLIENT_ID or not SP_CLIENT_SECRET:
        print("  SP-API credentials not configured, skipping global sales.")
        return {}, set()

    print("Loading global SKU mapping...")
    sku_map = load_global_sku_mapping()

    sales_raw   = {}
    unmapped    = set()
    now         = datetime.now(timezone.utc)
    cutoffs     = {
        "d7":  now - timedelta(days=7),
        "d15": now - timedelta(days=15),
        "d30": now - timedelta(days=30),
        "d90": now - timedelta(days=90),
    }

    for market_name, marketplace in MARKETPLACES.items():
        if not marketplace["refresh_token"]:
            print(f"  {market_name}: No refresh token, skipping.")
            continue

        print(f"  Fetching {market_name} orders...")
        try:
            orders, access_token = sp_get_orders(marketplace, days=days)
        except Exception as e:
            print(f"  {market_name}: Error: {e}")
            continue

        # Filter cancelled before fetching items
        valid_orders = [o for o in orders if o.get("OrderStatus") != "Canceled"]
        print(f"  {market_name}: {len(valid_orders)} valid orders. Fetching items (~{len(valid_orders)//60}min)...")
        start_time = time.time()

        for i, order in enumerate(valid_orders):
            if i % 50 == 0 and i > 0:
                elapsed = time.time() - start_time
                rate    = i / elapsed
                remaining = (len(valid_orders) - i) / rate if rate > 0 else 0
                print(f"    {market_name}: {i}/{len(valid_orders)} ({remaining/60:.1f}min remaining)...")

            order_id   = order.get("AmazonOrderId")
            order_date = order.get("PurchaseDate", "")

            try:
                order_dt = datetime.strptime(order_date, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
            except:
                continue

            items = sp_get_order_items(marketplace, order_id, access_token)
            time.sleep(0.05)  # 20 req/sec for order items API

            for item in items:
                global_sku = item.get("SellerSKU", "")
                qty        = int(item.get("QuantityOrdered", 0))
                if not global_sku or qty == 0:
                    continue

                master_sku = sku_map.get(global_sku)
                if not master_sku:
                    unmapped.add(global_sku)
                    continue

                if master_sku not in sales_raw:
                    sales_raw[master_sku] = []
                sales_raw[master_sku].append((order_dt, qty))

        print(f"  {market_name}: Done.")

    # Aggregate into time windows
    sales_by_master = {}
    for master_sku, entries in sales_raw.items():
        sales_by_master[master_sku] = {
            "d7":  sum(q for dt, q in entries if dt >= cutoffs["d7"]),
            "d15": sum(q for dt, q in entries if dt >= cutoffs["d15"]),
            "d30": sum(q for dt, q in entries if dt >= cutoffs["d30"]),
            "d90": sum(q for dt, q in entries),
        }

    if unmapped:
        print(f"\n  ⚠️ {len(unmapped)} Global SKUs not mapped to Master SKUs:")
        for s in sorted(unmapped):
            print(f"     {s}")
        print(f"  → Add these to data/global_sku_mapping.xlsx\n")

    print(f"  Global sales mapped: {len(sales_by_master)} master SKUs with global orders.")
    return sales_by_master, unmapped

# ── Fetch all inventory products ─────────────────────────────────────────────
def get_all_products():
    print("Fetching inventory products...")
    page = 1
    all_products = {}
    while True:
        data = bl_call("getInventoryProductsList", {"inventory_id": INVENTORY_ID, "page": page})
        products = data.get("products", {})
        if not products:
            break
        all_products.update(products)
        if len(products) < 1000:
            break
        page += 1
    print(f"  {len(all_products)} products found.")
    return all_products

# ── Fetch stock levels ────────────────────────────────────────────────────────
def get_stock(product_ids):
    print("Fetching stock levels...")
    BATCH = 200
    all_stock = {}
    for i in range(0, len(product_ids), BATCH):
        batch = product_ids[i:i+BATCH]
        data = bl_call("getInventoryProductsData", {
            "inventory_id": INVENTORY_ID,
            "products": batch
        })
        all_stock.update(data.get("products", {}))
        if i + BATCH < len(product_ids):
            time.sleep(0.15)
    print(f"  Stock fetched for {len(all_stock)} products.")
    return all_stock

def extract_stock(stock_info):
    if not stock_info:
        return 0
    warehouses = stock_info.get("stock", {})
    for key in [f"bl_{WAREHOUSE_ID}", str(WAREHOUSE_ID), WAREHOUSE_ID]:
        if str(key) in warehouses:
            return int(warehouses[str(key)] or 0)
    return 0

def extract_cost(stock_info):
    if not stock_info:
        return 0
    return float(stock_info.get("average_cost") or 0)

# ── Build bundle map ──────────────────────────────────────────────────────────
def build_bundle_map(products, stock_data):
    """
    Returns {bundle_sku: {component_sku: quantity}}
    e.g. {'SC+TC-LOTUS-SET-11': {'SC-LOTUS': 10, 'TC-TULIP-RUST': 1}}
    """
    # Build product_id → sku lookup
    id_to_sku = {}
    for pid_str, product in products.items():
        sku = product.get("sku") or pid_str
        id_to_sku[pid_str] = sku

    bundle_map = {}
    for pid_str, product in products.items():
        stock_info    = stock_data.get(pid_str, {})
        bundle_prods  = stock_info.get("bundle_products")
        if not bundle_prods:
            continue

        bundle_sku = product.get("sku") or pid_str
        components = {}
        for component_id_str, qty in bundle_prods.items():
            component_sku = id_to_sku.get(str(component_id_str))
            if component_sku:
                components[component_sku] = qty
        if components:
            bundle_map[bundle_sku] = components

    print(f"  Bundle map built: {len(bundle_map)} bundles with component SKUs.")
    print(f"  Product ID→SKU map: {len(id_to_sku)} entries.")
    return bundle_map, id_to_sku

# ── Fetch all orders since a date ─────────────────────────────────────────────
def get_orders_since(date_from):
    print(f"Fetching orders since {datetime.fromtimestamp(date_from).strftime('%Y-%m-%d')}...")
    all_orders = []
    date_confirmed_from = date_from

    # Pass 1: fetch confirmed orders (all regular platforms) via date_confirmed pagination
    while True:
        data = bl_call("getOrders", {
            "date_confirmed_from": date_confirmed_from,
            "get_unconfirmed_orders": False
        })
        orders = data.get("orders", [])
        if not orders:
            break
        all_orders.extend(orders)
        print(f"  Fetched {len(all_orders)} confirmed orders so far...")
        if len(orders) < 100:
            break
        last_date = orders[-1].get("date_confirmed", 0)
        if not last_date or last_date <= date_confirmed_from:
            break
        date_confirmed_from = last_date + 1
        time.sleep(0.15)

    confirmed_count = len(all_orders)
    print(f"  Confirmed orders: {confirmed_count}")

    # Pass 2: fetch unconfirmed Vendor DF orders (date_confirmed = 0)
    # Must paginate using id_from since date_confirmed_from doesn't work for these
    existing_ids = {o["order_id"] for o in all_orders}

    for source_id in [445, 537]:
        vendor_count = 0
        id_from = 0

        while True:
            params = {
                "date_from": date_from,
                "get_unconfirmed_orders": True,
                "filter_order_source": "amazon_vendor",
                "filter_order_source_id": source_id,
            }
            if id_from:
                params["id_from"] = id_from

            data = bl_call("getOrders", params)
            vendor_orders = data.get("orders", [])
            if not vendor_orders:
                break

            # Filter by date_add and exclude already fetched
            new_orders = [
                o for o in vendor_orders
                if o.get("date_add", 0) >= date_from
                and o["order_id"] not in existing_ids
            ]
            all_orders.extend(new_orders)
            existing_ids.update(o["order_id"] for o in new_orders)
            vendor_count += len(new_orders)

            if len(vendor_orders) < 100:
                break

            id_from = vendor_orders[-1].get("order_id", 0)
            if not id_from:
                break
            time.sleep(0.15)

        if vendor_count:
            print(f"  Added {vendor_count} Amazon Vendor DF orders (source {source_id})")

    print(f"  Total {len(all_orders)} orders fetched.")
    return all_orders

# ── Fetch all returns since a date ────────────────────────────────────────────
def get_returns_since(date_from):
    print(f"Fetching returns since {datetime.fromtimestamp(date_from).strftime('%Y-%m-%d')}...")
    all_returns = []
    last_return_id = 0
    max_pages = 500  # safety cap — 500 × 100 = 50,000 returns max

    for page in range(max_pages):
        try:
            params = {"date_from": date_from}
            if last_return_id:
                params["id_from"] = last_return_id + 1  # BL pagination requires +1

            data = bl_call("getOrderReturns", params)
            returns = data.get("returns", [])
            if not returns:
                break

            all_returns.extend(returns)

            if len(all_returns) % 1000 == 0:
                print(f"  Fetched {len(all_returns)} returns so far...")

            if len(returns) < 100:
                break

            new_last_id = returns[-1].get("return_id", 0)
            if not new_last_id:
                break
            last_return_id = new_last_id
            time.sleep(0.15)

        except Exception as e:
            print(f"  Returns fetch error on page {page+1}: {e}")
            break

    print(f"  Total {len(all_returns)} returns fetched.")
    return all_returns

# ── Aggregate sales per SKU with bundle remapping ────────────────────────────
def aggregate_sales(orders, bundle_map={}, id_to_sku={}):
    sales = {}
    for order in orders:
        if order.get("order_source") == "order_return":
            continue
        # Skip cancelled orders
        if order.get("order_status_id") == CANCELLED_STATUS_ID:
            continue
        platform = get_platform(order.get("order_source_id", 0))
        for product in order.get("products", []):
            product_id = str(product.get("product_id", ""))
            sku = id_to_sku.get(product_id) or product.get("sku", "")
            if not sku:
                continue
            qty = int(product.get("quantity", 0))

            if sku in bundle_map:
                for component_sku, component_qty in bundle_map[sku].items():
                    total_qty = qty * component_qty
                    if component_sku not in sales:
                        sales[component_sku] = {"total": 0, "amazon": 0, "amazon_vendor": 0,
                                                "flipkart": 0, "myntra": 0, "ajio": 0, "others": 0}
                    sales[component_sku]["total"] += total_qty
                    sales[component_sku][platform] = sales[component_sku].get(platform, 0) + total_qty
            else:
                if sku not in sales:
                    sales[sku] = {"total": 0, "amazon": 0, "amazon_vendor": 0,
                                  "flipkart": 0, "myntra": 0, "ajio": 0, "others": 0}
                sales[sku]["total"] += qty
                sales[sku][platform] = sales[sku].get(platform, 0) + qty
    return sales

def aggregate_returns(orders, id_to_sku={}):
    """Returns from orders panel (order_source = order_return)"""
    returns = {}
    for order in orders:
        if order.get("order_source") != "order_return":
            continue
        for product in order.get("products", []):
            product_id = str(product.get("product_id", ""))
            sku = id_to_sku.get(product_id) or product.get("sku", "")
            if not sku:
                continue
            qty = int(product.get("quantity", 0))
            returns[sku] = returns.get(sku, 0) + qty
    return returns

def aggregate_returns_panel(panel_returns, id_to_sku={}):
    """Returns from Returns panel via getOrderReturns — completed returns only"""
    returns = {}
    for ret in panel_returns:
        for product in ret.get("products", []):
            product_id = str(product.get("product_id", ""))
            sku = id_to_sku.get(product_id) or product.get("sku", "")
            if not sku:
                continue
            qty = int(product.get("quantity", 0))
            returns[sku] = returns.get(sku, 0) + qty
    return returns

# ── Velocity calculation ──────────────────────────────────────────────────────
def calc_velocity(net_sales):
    avgs = {
        "d7":  net_sales["d7"]  / 7,
        "d15": net_sales["d15"] / 15,
        "d30": net_sales["d30"] / 30,
        "d90": net_sales["d90"] / 90,
    }
    has_data   = {k: net_sales[k] > 0 for k in ["d7", "d15", "d30", "d90"]}
    total_weight = sum(WEIGHTS[k] for k in WEIGHTS if has_data[k])
    if total_weight == 0:
        return 0.0

    weighted = sum(avgs[k] * WEIGHTS[k] for k in WEIGHTS if has_data[k])
    return max(0.0, weighted / total_weight)

def calc_metrics(current_stock, net_sales):
    velocity       = calc_velocity(net_sales)
    days_remaining = (current_stock / velocity) if velocity > 0 else None
    suggested_order = max(0, round((TARGET_DAYS - (days_remaining or 0)) * velocity)) if velocity > 0 else 0

    # Status hierarchy:
    # Dead Stock   → velocity < 0.1 AND stock > 0 (overrides all)
    # Out of Stock → stock = 0
    # Critical     → days ≤ 7
    # Low          → days ≤ 15
    # Watch        → days ≤ 30
    # Healthy      → 30 < days ≤ 365
    # Overstock    → days > 365

    DEAD_VELOCITY_THRESHOLD = 0.1

    if current_stock <= 0:
        status = "OUT_OF_STOCK"
    elif velocity < DEAD_VELOCITY_THRESHOLD:
        status = "DEAD_STOCK"
    elif days_remaining is None:
        status = "HEALTHY"
    elif days_remaining <= 7:
        status = "CRITICAL"
    elif days_remaining <= 15:
        status = "LOW"
    elif days_remaining <= 30:
        status = "WATCH"
    elif days_remaining <= 365:
        status = "HEALTHY"
    else:
        status = "OVERSTOCK"

    return {
        "dailyVelocity":  round(velocity, 3),
        "daysRemaining":  round(days_remaining, 1) if days_remaining is not None else None,
        "suggestedOrder": suggested_order,
        "status":         status,
    }

# ── Build full report ─────────────────────────────────────────────────────────
def build_report():
    products    = get_all_products()
    product_ids = [int(k) for k in products.keys()]
    stock_data  = get_stock(product_ids)

    print("Building bundle map...")
    bundle_map, id_to_sku = build_bundle_map(products, stock_data)

    # Always fetch full 90 days fresh — ensures accurate cancelled/RTO status
    from90     = days_ago(90)
    all_orders = get_orders_since(from90)

    # Build cancelled order ID set for return filtering
    cancelled_order_ids = {
        o["order_id"] for o in all_orders
        if o.get("order_status_id") == CANCELLED_STATUS_ID
    }
    print(f"  Cancelled orders in 90d: {len(cancelled_order_ids)}")

    orders90 = [o for o in all_orders if (o.get("date_confirmed") or o.get("date_add", 0)) >= from90]

    # Fetch returns from Returns panel
    print("Fetching returns from Returns panel...")
    all_panel_returns = get_returns_since(from90)
    REJECTED_STATUS_ID = 5829
    valid_returns = [r for r in all_panel_returns if r.get("status_id") != REJECTED_STATUS_ID]
    panel_returns = [r for r in valid_returns if r.get("order_id") not in cancelled_order_ids]
    print(f"  Valid returns (excl. rejected + RTO): {len(panel_returns)} of {len(all_panel_returns)}")

    # Fetch global sales from Amazon SP-API
    print("\nFetching Global Sales (US/UK/UAE)...")
    global_sales, unmapped_skus = get_global_sales(days=90)
    print(f"  Global sales: {len(global_sales)} master SKUs\n")

    ts = {"d7": days_ago(7), "d15": days_ago(15), "d30": days_ago(30)}

    def filter_orders(days_key):
        return [o for o in orders90 if (o.get("date_confirmed") or o.get("date_add", 0)) >= ts[days_key]]

    def filter_returns(days_key):
        return [r for r in panel_returns if r.get("date_add", 0) >= ts[days_key]]

    orders_d7  = filter_orders("d7")
    orders_d15 = filter_orders("d15")
    orders_d30 = filter_orders("d30")

    panel_returns_d7  = filter_returns("d7")
    panel_returns_d15 = filter_returns("d15")
    panel_returns_d30 = filter_returns("d30")

    sales = {
        "d7":  aggregate_sales(orders_d7,  bundle_map, id_to_sku),
        "d15": aggregate_sales(orders_d15, bundle_map, id_to_sku),
        "d30": aggregate_sales(orders_d30, bundle_map, id_to_sku),
        "d90": aggregate_sales(orders90,   bundle_map, id_to_sku),
    }

    # Merge global sales into BL sales
    for window in ["d7", "d15", "d30", "d90"]:
        for master_sku, global_window in global_sales.items():
            qty = global_window.get(window, 0)
            if qty == 0:
                continue
            if master_sku not in sales[window]:
                sales[window][master_sku] = {"total": 0, "amazon": 0, "amazon_vendor": 0,
                                              "flipkart": 0, "myntra": 0, "ajio": 0,
                                              "others": 0, "global": 0}
            sales[window][master_sku]["total"]  = sales[window][master_sku].get("total", 0) + qty
            sales[window][master_sku]["global"] = sales[window][master_sku].get("global", 0) + qty

    # Merge returns from orders panel + returns panel
    def merge_returns(orders_ret, panel_ret):
        r1 = aggregate_returns(orders_ret, id_to_sku)
        r2 = aggregate_returns_panel(panel_ret, id_to_sku)
        merged = dict(r1)
        for sku, qty in r2.items():
            merged[sku] = merged.get(sku, 0) + qty
        return merged

    returns = {
        "d7":  merge_returns(orders_d7,  panel_returns_d7),
        "d15": merge_returns(orders_d15, panel_returns_d15),
        "d30": merge_returns(orders_d30, panel_returns_d30),
        "d90": merge_returns(orders90,   panel_returns),
    }

    report = []
    for product_id_str, product in products.items():
        stock_info = stock_data.get(product_id_str, {})

        # Skip bundle products
        if stock_info.get("bundle_products"):
            continue

        sku = product.get("sku") or product_id_str

        # Skip explicitly excluded SKUs
        if sku in EXCLUDED_SKUS:
            continue

        product_id    = int(product_id_str)
        name          = product.get("name") or "Unknown"
        current_stock = extract_stock(stock_data.get(product_id_str))
        avg_cost      = extract_cost(stock_data.get(product_id_str))
        stock_value   = round(current_stock * avg_cost)

        net_sales = {
            "d7":  max(0, sales["d7"].get(sku, {}).get("total", 0)  - returns["d7"].get(sku, 0)),
            "d15": max(0, sales["d15"].get(sku, {}).get("total", 0) - returns["d15"].get(sku, 0)),
            "d30": max(0, sales["d30"].get(sku, {}).get("total", 0) - returns["d30"].get(sku, 0)),
            "d90": max(0, sales["d90"].get(sku, {}).get("total", 0) - returns["d90"].get(sku, 0)),
        }

        platform_sales = {
            "amazon":        sales["d30"].get(sku, {}).get("amazon", 0),
            "amazon_vendor": sales["d30"].get(sku, {}).get("amazon_vendor", 0),
            "flipkart":      sales["d30"].get(sku, {}).get("flipkart", 0),
            "myntra":        sales["d30"].get(sku, {}).get("myntra", 0),
            "ajio":          sales["d30"].get(sku, {}).get("ajio", 0),
            "others":        sales["d30"].get(sku, {}).get("others", 0),
            "global":        sales["d30"].get(sku, {}).get("global", 0),
        }

        metrics    = calc_metrics(current_stock, net_sales)
        dead_stock = metrics["status"] == "DEAD_STOCK"
        adj_order  = round(metrics["suggestedOrder"] * SEASONAL_MULT)

        report.append({
            "productId":    product_id,
            "sku":          sku,
            "name":         name,
            "currentStock": current_stock,
            "netSales":     net_sales,
            "grossSales": {
                "d7":  sales["d7"].get(sku, {}).get("total", 0),
                "d15": sales["d15"].get(sku, {}).get("total", 0),
                "d30": sales["d30"].get(sku, {}).get("total", 0),
                "d90": sales["d90"].get(sku, {}).get("total", 0),
            },
            "returns": {
                "d7":  returns["d7"].get(sku, 0),
                "d15": returns["d15"].get(sku, 0),
                "d30": returns["d30"].get(sku, 0),
                "d90": returns["d90"].get(sku, 0),
            },
            "platformSales":  platform_sales,
            "dailyVelocity":  metrics["dailyVelocity"],
            "daysRemaining":  metrics["daysRemaining"],
            "suggestedOrder": adj_order,
            "status":         metrics["status"],
            "isDeadStock":    dead_stock,
            "avgCost":        avg_cost,
            "stockValue":     stock_value,
        })

    status_order = {"OUT_OF_STOCK": 0, "CRITICAL": 1, "LOW": 2, "WATCH": 3, "HEALTHY": 4, "OVERSTOCK": 5, "DEAD_STOCK": 6}
    report.sort(key=lambda p: status_order.get(p["status"], 5))

    summary = {
        "total":            len(report),
        "outOfStock":       sum(1 for p in report if p["status"] == "OUT_OF_STOCK"),
        "critical":         sum(1 for p in report if p["status"] == "CRITICAL"),
        "low":              sum(1 for p in report if p["status"] == "LOW"),
        "watch":            sum(1 for p in report if p["status"] == "WATCH"),
        "healthy":          sum(1 for p in report if p["status"] == "HEALTHY"),
        "overstock":        sum(1 for p in report if p["status"] == "OVERSTOCK"),
        "deadStock":        sum(1 for p in report if p["status"] == "DEAD_STOCK"),
        "totalStockValue":   sum(p["stockValue"] for p in report),
        "deadStockValue":    sum(p["stockValue"] for p in report if p["status"] == "DEAD_STOCK"),
        "healthyStockValue": sum(p["stockValue"] for p in report if p["status"] == "HEALTHY"),
        "overstockValue":    sum(p["stockValue"] for p in report if p["status"] == "OVERSTOCK"),
        "globalSKUsCount":   len(global_sales),
        "unmappedGlobalSKUs": sorted(list(unmapped_skus)),
        "generatedAt":      datetime.now(timezone.utc).isoformat(),
        "seasonalMult":     SEASONAL_MULT,
    }

    return {"summary": summary, "products": report}

# ── Upload to GitHub ──────────────────────────────────────────────────────────
def upload_to_github(data):
    print("Uploading report to GitHub...")
    # Debug: show first 10 chars of token to verify correct one is being used
    token_preview = GITHUB_TOKEN[:10] + "..." if GITHUB_TOKEN and len(GITHUB_TOKEN) > 10 else "EMPTY/MISSING"
    print(f"  Using token: {token_preview}")
    content = json.dumps(data, ensure_ascii=False)
    encoded = base64.b64encode(content.encode()).decode()

    api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"

    sha = None
    try:
        req = urllib.request.Request(api_url, headers={
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json",
        })
        with urllib.request.urlopen(req) as resp:
            existing = json.loads(resp.read().decode())
            sha = existing.get("sha")
    except:
        pass

    payload = {
        "message": f"Daily stock report {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "content": encoded,
    }
    if sha:
        payload["sha"] = sha

    req = urllib.request.Request(
        api_url,
        data=json.dumps(payload).encode(),
        headers={
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json",
        },
        method="PUT"
    )
    with urllib.request.urlopen(req) as resp:
        print(f"  Uploaded successfully. Status: {resp.status}")

# ── Save velocity snapshot ────────────────────────────────────────────────────
def save_snapshot(products):
    today      = datetime.now().strftime("%Y-%m-%d")
    file_exists = SNAPSHOT_CSV.exists()
    with open(SNAPSHOT_CSV, "a", newline="", encoding="utf-8") as f:
        fieldnames = ["date","sku","name","currentStock","dailyVelocity",
                      "netSales7d","netSales15d","netSales30d","netSales90d",
                      "status","daysRemaining","suggestedOrder"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        for p in products:
            writer.writerow({
                "date":          today,
                "sku":           p["sku"],
                "name":          p["name"],
                "currentStock":  p["currentStock"],
                "dailyVelocity": p["dailyVelocity"],
                "netSales7d":    p["netSales"]["d7"],
                "netSales15d":   p["netSales"]["d15"],
                "netSales30d":   p["netSales"]["d30"],
                "netSales90d":   p["netSales"]["d90"],
                "status":        p["status"],
                "daysRemaining": p["daysRemaining"] or "",
                "suggestedOrder":p["suggestedOrder"],
            })
    print(f"  Snapshot saved ({len(products)} rows).")

def save_value_history(summary):
    """Save daily total/healthy/dead stock values for trend graph."""
    today      = datetime.now().strftime("%Y-%m-%d")
    file_exists = VALUE_HISTORY_CSV.exists()

    # Check if today already logged (avoid duplicates if run twice)
    if file_exists:
        with open(VALUE_HISTORY_CSV, "r", encoding="utf-8") as f:
            last_line = f.readlines()[-1] if f.readlines() else ""
        if last_line.startswith(today):
            print(f"  Value history already logged for {today}, skipping.")
            return

    with open(VALUE_HISTORY_CSV, "a", newline="", encoding="utf-8") as f:
        fieldnames = ["date","totalStockValue","healthyStockValue","deadStockValue",
                      "totalSKUs","healthySKUs","deadSKUs","oosSKUs"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerow({
            "date":             today,
            "totalStockValue":  summary.get("totalStockValue", 0),
            "healthyStockValue":summary.get("healthyStockValue", 0),
            "deadStockValue":   summary.get("deadStockValue", 0),
            "totalSKUs":        summary.get("total", 0),
            "healthySKUs":      summary.get("healthy", 0),
            "deadSKUs":         summary.get("deadStock", 0),
            "oosSKUs":          summary.get("outOfStock", 0),
        })
    print(f"  Value history saved for {today}.")

# ── Send HTML email ───────────────────────────────────────────────────────────
def send_email(data):
    summary  = data["summary"]
    products = data["products"]
    now      = datetime.now().strftime("%d %b %Y, %I:%M %p IST")

    urgent = sorted([p for p in products if p["status"] in ("OUT_OF_STOCK","CRITICAL")], key=lambda x: x["suggestedOrder"], reverse=True)
    low    = sorted([p for p in products if p["status"] in ("LOW","WATCH")], key=lambda x: x["suggestedOrder"], reverse=True)
    dead   = [p for p in products if p["status"] == "DEAD_STOCK"]
    over   = [p for p in products if p["status"] == "OVERSTOCK"]

    STATUS_COLOR = {
        "OUT_OF_STOCK": "#e05c5c", "CRITICAL": "#f07830",
        "LOW": "#f0a500", "WATCH": "#a0a0ff", "HEALTHY": "#5ce0a0"
    }
    STATUS_EMOJI = {
        "OUT_OF_STOCK": "🔴", "CRITICAL": "🟠",
        "LOW": "🟡", "WATCH": "🔵", "HEALTHY": "🟢"
    }

    def platform_cell(p):
        ps    = p.get("platformSales", {})
        parts = []
        if ps.get("amazon",0)>0:        parts.append(f'<span style="color:#ff9900">AM:{ps["amazon"]}</span>')
        if ps.get("flipkart",0)>0:      parts.append(f'<span style="color:#2f74ff">FL:{ps["flipkart"]}</span>')
        if ps.get("myntra",0)>0:        parts.append(f'<span style="color:#ff3f6c">MY:{ps["myntra"]}</span>')
        if ps.get("ajio",0)>0:          parts.append(f'<span style="color:#00c896">AJ:{ps["ajio"]}</span>')
        if ps.get("global",0)>0:        parts.append(f'<span style="color:#b478ff">GL:{ps["global"]}</span>')
        if ps.get("others",0)>0:        parts.append(f'<span style="color:#a0a0a0">OT:{ps["others"]}</span>')
        return " ".join(parts) or "—"

    # Unmapped global SKUs warning
    unmapped = summary.get("unmappedGlobalSKUs", [])
    unmapped_note = ""
    if unmapped:
        unmapped_note = f'<div style="background:rgba(240,165,0,0.1);border:1px solid rgba(240,165,0,0.3);border-radius:8px;padding:12px 16px;margin-bottom:20px;color:#f0a500">⚠ <strong>{len(unmapped)} Global SKUs not mapped</strong> → add to data/global_sku_mapping.xlsx:<br><span style="font-family:monospace;font-size:12px;color:#ccc">{", ".join(unmapped[:10])}{"..." if len(unmapped) > 10 else ""}</span></div>'

    def product_rows(plist, limit=50):
        rows = ""
        for p in plist[:limit]:
            color       = STATUS_COLOR.get(p["status"], "#999")
            emoji       = STATUS_EMOJI.get(p["status"], "⚪")
            days        = f'{p["daysRemaining"]}d' if p["daysRemaining"] is not None else "∞"
            order_color = "#e05c5c" if p["status"] in ("OUT_OF_STOCK","CRITICAL") else "#f0a500"
            rows += f"""
            <tr style="border-bottom:1px solid #1e1e2e">
              <td style="padding:10px 14px;color:{color};font-weight:700">{emoji} {p["status"].replace("_"," ")}</td>
              <td style="padding:10px 14px;font-family:monospace;font-size:12px;color:#888">{p["sku"]}</td>
              <td style="padding:10px 14px">{p["name"][:35]}</td>
              <td style="padding:10px 14px;text-align:right;font-family:monospace">{p["currentStock"]}</td>
              <td style="padding:10px 14px;text-align:right;color:{color};font-family:monospace">{days}</td>
              <td style="padding:10px 14px;text-align:right;font-family:monospace">{p["dailyVelocity"]}</td>
              <td style="padding:10px 14px;text-align:right;font-family:monospace">{p["netSales"]["d7"]}</td>
              <td style="padding:10px 14px;text-align:right;font-family:monospace">{p["netSales"]["d30"]}</td>
              <td style="padding:10px 14px;font-size:11px">{platform_cell(p)}</td>
              <td style="padding:10px 14px;text-align:right;font-weight:700;color:{order_color}">{p["suggestedOrder"] or "—"}</td>
            </tr>"""
        return rows

    def make_table(rows):
        return f"""<table style="width:100%;border-collapse:collapse;font-size:13px;background:#111118;border-radius:10px;overflow:hidden">
        <thead><tr style="background:#1a1a26;font-size:11px;text-transform:uppercase;letter-spacing:1px;color:#5a5a7a">
          <th style="padding:10px 14px;text-align:left">Status</th>
          <th style="padding:10px 14px;text-align:left">SKU</th>
          <th style="padding:10px 14px;text-align:left">Product</th>
          <th style="padding:10px 14px;text-align:right">Stock</th>
          <th style="padding:10px 14px;text-align:right">Days</th>
          <th style="padding:10px 14px;text-align:right">Vel/d</th>
          <th style="padding:10px 14px;text-align:right">7d</th>
          <th style="padding:10px 14px;text-align:right">30d</th>
          <th style="padding:10px 14px;text-align:left">Platforms</th>
          <th style="padding:10px 14px;text-align:right">Order</th>
        </tr></thead>
        <tbody>{rows}</tbody></table>"""

    seasonal_note = f'<p style="color:#f0a500;margin:8px 0 0">⚡ Seasonal {SEASONAL_MULT}× applied</p>' if SEASONAL_MULT != 1.0 else ""

    html = f"""<!DOCTYPE html><html><body style="background:#0a0a0f;color:#f0f0f8;font-family:'Segoe UI',sans-serif;margin:0;padding:0">
<div style="max-width:900px;margin:0 auto;padding:32px 24px">
  <div style="margin-bottom:32px">
    <div style="font-size:24px;font-weight:800">Weavers Villa</div>
    <div style="font-size:11px;letter-spacing:3px;text-transform:uppercase;color:#f0a500;margin-top:4px">Daily Stock Intelligence — {now}</div>
    {seasonal_note}
  </div>
  {unmapped_note}
  <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:32px">
    {"".join(f'<div style="background:#111118;border:1px solid #2a2a3a;border-radius:10px;padding:16px 20px;min-width:110px"><div style="font-size:32px;font-weight:800;color:{c}">{v}</div><div style="font-size:11px;text-transform:uppercase;letter-spacing:1px;color:#666;margin-top:4px">{l}</div></div>'
    for c,v,l in [
        ("#e05c5c", summary["outOfStock"],  "Out of Stock"),
        ("#f07830", summary["critical"],    "Critical <7d"),
        ("#f0a500", summary["low"],         "Low <15d"),
        ("#5ce0a0", summary["healthy"],     "Healthy"),
        ("#5a5a7a", summary["deadStock"],   "Dead Stock"),
    ])}
  </div>
  {'<div style="margin-bottom:32px"><div style="font-size:14px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#e05c5c;margin-bottom:12px">🔴 Urgent — Order Immediately</div>' + make_table(product_rows(urgent)) + '</div>' if urgent else ''}
  {'<div style="margin-bottom:32px"><div style="font-size:14px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#f0a500;margin-bottom:12px">🟡 Low & Watch — Plan Orders</div>' + make_table(product_rows(low)) + '</div>' if low else ''}
  {'<div style="margin-bottom:32px"><div style="font-size:14px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#5a5a7a;margin-bottom:12px">💀 Dead Stock — Zero Sales 30d</div><table style="width:100%;border-collapse:collapse;font-size:13px;background:#111118;border-radius:10px;overflow:hidden"><thead><tr style="background:#1a1a26;font-size:11px;text-transform:uppercase;letter-spacing:1px;color:#5a5a7a"><th style="padding:10px 14px;text-align:left">SKU</th><th style="padding:10px 14px;text-align:left">Product</th><th style="padding:10px 14px;text-align:right">Stock</th></tr></thead><tbody>' + "".join(f'<tr style="border-bottom:1px solid #1e1e2e;opacity:0.6"><td style="padding:10px 14px;font-family:monospace;font-size:12px;color:#5a5a7a">{p["sku"]}</td><td style="padding:10px 14px">{p["name"][:40]}</td><td style="padding:10px 14px;text-align:right;font-family:monospace">{p["currentStock"]}</td></tr>' for p in dead[:30]) + '</tbody></table></div>' if dead else ''}
  <div style="font-size:12px;color:#333;margin-top:32px;border-top:1px solid #1a1a2a;padding-top:16px">
    Generated by Weavers Villa Stock Intelligence · {summary["total"]} SKUs · <a href="https://weavers-stock.vercel.app" style="color:#f0a500">weavers-stock.vercel.app</a>
  </div>
</div></body></html>"""

    urgent_count = summary["outOfStock"] + summary["critical"]
    subject = f"🔴 {urgent_count} Urgent | 🟡 {summary['low']} Low | Weavers Stock {datetime.now().strftime('%d %b')}" if urgent_count else f"✅ Stock OK | 🟡 {summary['low']} Low | Weavers Stock {datetime.now().strftime('%d %b')}"

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = GMAIL_FROM
    msg["To"]      = GMAIL_TO
    msg.attach(MIMEText(html, "html"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_FROM, GMAIL_APP_PASS)
        server.sendmail(GMAIL_FROM, GMAIL_TO, msg.as_string())
    print(f"  Email sent to {GMAIL_TO}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"\n{'='*60}")
    print(f"Weavers Villa Stock Report — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    data    = build_report()
    summary = data["summary"]
    print(f"\nSummary: {summary['total']} SKUs | OOS:{summary['outOfStock']} Critical:{summary['critical']} Low:{summary['low']} Dead:{summary['deadStock']}")

    print("\nSaving snapshot...")
    save_snapshot(data["products"])
    save_value_history(data["summary"])

    print("\nUploading to GitHub...")
    upload_to_github(data)

    print("\nSending email...")
    send_email(data)

    print(f"\n✅ Done! Dashboard: https://weavers-stock.vercel.app")

if __name__ == "__main__":
    main()
